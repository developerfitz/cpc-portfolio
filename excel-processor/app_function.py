from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, get_column_interval
from openpyxl.styles import Alignment, Border, Font
import requests
import boto3
from aws_xray_sdk.core import xray_recorder
from aws_xray_sdk.core import patch_all

import json
import logging
import os
from classes import Symbol, Examiner


logger = logging.getLogger()
logger.setLevel(logging.INFO)
patch_all()

s3 = boto3.client('s3')
BUCKET = os.environ['BUCKET']
PRE = os.environ['PRE']
POST = os.environ['POST']


def excel_processor(event, context):
    # TODO: need some error handling in here
    # download excel sheet to work on
    logger.info(BUCKET)
    # TODO: should add the filename from the event object
    # logger.info(event)
    body = json.loads(event['body'])
    # logger.info(body)
    filename = body['filename']
    logger.info(filename)
    
    # s3.download_file(BUCKET, f'{PRE}/{filename}', f'/tmp/{filename}')
    # BUCKET/private/${cognito-id}/PRE/filename 
    # ??: write file 
    wb = load_workbook(filename=f'/tmp/{filename}', read_only=True)

    # TODO: take in the name of the sheet that is in the file
    sheet = wb["Symbols Sorted"]
    examiner = Examiner(first_name="dark", last_name='examiner')
    portfolio_list = []
    # portfolio_data = sheet["B1:H673"]

    # first row labels of ws
    SYMBOL = 0
    C_STAR = 3
    TALLY = 4
    QUALIFIED = 6
    ROWS_END = 673

    for row in sheet.iter_rows(
        min_row=2,
        max_row=ROWS_END,
        min_col=2,
        max_col=8,
        values_only=True):
        if row[SYMBOL]:
            portfolio_symbol = Symbol(
                symbol=row[SYMBOL],
                c_star=row[C_STAR],
                tally=row[TALLY],
                qualified=row[QUALIFIED]
                )
            # adding symbol only to a list
            portfolio_list.append(portfolio_symbol.symbol)

            # adding symbol object to examiner object as list
            examiner.portfolio.append(portfolio_symbol)

            # adding symbol object to examiner object as dict
            row_symbol = row[SYMBOL]
            examiner.symbols[row_symbol] = portfolio_symbol

    # provide a list of symbols only on examiner for easy processing
    examiner.symbols_list_only = portfolio_list
    logger.info(examiner.symbols_list_only[0:5])


    # # TODO: possibly turn this into a stream, so the limit does not matter
    PATENTVIEW = 'https://api.patentsview.org/cpc_subsections/query'
    payload = {
      "q": {"cpc_subgroup_id": examiner.symbols_list_only[0:5]},
      "f": ["cpc_subgroup_id", "cpc_subgroup_title"],
      "s": [{"cpc_subgroup_id": "asc"}],
      "o": {"matched_subentities_only": 'true'}
    }
    logger.info(payload)
    req = requests.post(PATENTVIEW, data=json.dumps(payload))
    if req.status_code != 200:
        logger.info(req.headers['x-status-reason'])
        logger.info(req.raise_for_status())
        # return(req.raise_for_status())   

    for subsection in req.json()['cpc_subsections']:
        for subgroup in subsection['cpc_subgroups']:
            subgroup_id = subgroup['cpc_subgroup_id']
            subgroup_title = subgroup['cpc_subgroup_title']
            examiner.symbols[subgroup_id].title = subgroup_title


    # make new workbook and send to user
    new_wb = Workbook()
    ws = new_wb.active
    ws.title = 'Processed Portfolio'
    column_labels = ['Symbol', 'Title', 'Tally', 'C*', 'Qualified']
    ws.append(column_labels)

    for symbol in examiner.portfolio:
        row_data = [symbol.symbol,
                    symbol.title,
                    symbol.tally,
                    symbol.c_star,
                    symbol.qualified
                    ]
        ws.append(row_data)

    symbol_cells = ws['A']
    title_cells = ws['B']
    c_star_cells = ws['C']
    tally_cells = ws['D']
    qualified_cells = ws['E']
    styles = {
      "align": Alignment(vertical='center', wrap_text=True),
      "align-center": Alignment(horizontal='center', vertical='center', wrap_text=True),
      "font": Font(size=16)
    }
    # TODO: add column_dimensions to adjust the width of a column
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 150
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    for cell in symbol_cells:
        cell.alignment = styles['align']
        cell.font = styles['font']
    for cell in title_cells:
        cell.alignment = styles['align']
        cell.font = styles['font']
    for cell in c_star_cells:
        cell.alignment = styles['align-center']
        cell.font = styles['font']
    for cell in tally_cells:
        cell.alignment = styles['align-center']
        cell.font = styles['font']
    for cell in qualified_cells:
        cell.alignment = styles['align-center']
        cell.font = styles['font']

    # save and upload processed excel wb
    processed_filename = f'processed-{filename}'
    key_after_processing = f'{POST}/{processed_filename}'
    tmp_file_location = f'/tmp/{processed_filename}'
    logger.info(processed_filename)
    new_wb.save(tmp_file_location)

    s3.upload_file(
        tmp_file_location,
        BUCKET,
        key_after_processing
        )

    s3_params = {
        'Bucket': BUCKET,
        'Key': key_after_processing,
    }

    logger.info(s3_params)
    url = s3.generate_presigned_url(
        'get_object',
        Params=s3_params,
        HttpMethod='GET',
        ExpiresIn=60
    )

    payload = {
        'message': 'Hope this helps you organize your portfolio',
        'bucket': BUCKET,
        'filename': filename,
        'presignedUrl': url,
    }
    logger.info(payload)

    return {
        'headers': {
            'Access-Control-Allow-Origin': 'http://localhost:8888',
        },
        'statusCode': 200,
        'body': json.dumps(payload),
    }


# def main():
    # handle incoming request
    # get necessary items from request

    # def parse_examiner_Workbook(wb, sheet):
        # def extract_examiner_data():
            # makes examiner object to store data from workbook
            # return examiner object
        # def get_CPC_titles(examiner_object):
            # API call to patentview for titles
            # return dict {subgroup_id: subgroup_title}
        
    # append the titles to the examiner object
    # create new workbook and added examiner data (symbols + title)
        
    # def style_workbook():
        # return styled worksheet
    # save workbook to tmp and prepare to upload to S3

    # def generate_presignedUrl(bucket, key):
        #return url
    
    # prepare payload
    # return payload in a response

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, get_column_interval
from openpyxl.styles import Alignment, Border, Font
import requests

import json
# from tempfile import NamedTemporaryFile
from classes import Symbol, Examiner

PORTFOLIO = "CPC_Portfolio_Determination-Fitzhugh_Julius.xlsx"
# workbook (wb)
wb = load_workbook(filename=PORTFOLIO, read_only=True)
# used for streams
# with NamedTemporaryFile() as tmp:
#   wb.save(tmp.name)
#   tmp.seek(0)
#   stream = tmp.read()

# TODO: take in the name of the sheet that is in the file
sheet = wb["Symbol Sorted"]

portfolio_data = sheet["B1:H673"]
# print(portfolio_data)

examiner = Examiner(first_name="julius", last_name='fitzhugh')
examiner_portfolio = []
# TODO: implement a parcer that takes in the row 1 that
# should be the the name of the columns
# (Symbol,	Manually added,	Added by,	C* designation,	tally total,	Increase tally,	Qualified)
SYMBOL = 0
C_STAR = 3
TALLY = 4
QUALIFIED = 6
ROWS_END = 673

for row in sheet.iter_rows(min_row=2, max_row=ROWS_END, min_col=2, max_col=8, values_only=True):
    if row[SYMBOL]:
        portfolio_symbol = Symbol(symbol=row[SYMBOL],
                                  c_star=row[C_STAR],
                                  tally=row[TALLY],
                                  qualified=row[QUALIFIED]
                                 )
        examiner_portfolio.append(portfolio_symbol.symbol)
        examiner.portfolio.append(portfolio_symbol)
        row_symbol = row[SYMBOL]
        examiner.symbols[row_symbol] = portfolio_symbol
        examiner.symbols_list_only = examiner_portfolio
        # print(examiner.symbols[row_symbol])
    # cpc_symbol = row[0]
    # }
    #   "symbol": row[0], #TODO: format to be readable
    #   "title": None,
    #   "c_star": row[3],
    #   "tallies": str(row[4])
    # }
    # examiner_portfolio[cpc_symbol] = portfolio_symbol
# print(json.dumps(examiner_portfolio))
# print(examiner.symbols_list_only)

# symbols_list = []
# for list_symbol in examiner.portfolio[0:5]:
#     print(list_symbol)
#     symbols_list.append(list_symbol)

# print('dictnary list')
# for dict_symbol in symbols_list:
#     print(examiner.symbols[dict_symbol.symbol])

# # print(dict(examiner.portfolio))
# for symbol in examiner.portfolio:
#     print(symbol)


# example 1
# https://www.patentsview.org/api/cpc_subsections/query?q={"cpc_subgroup_id":"B32B1/06"}&f=["cpc_subgroup_id","cpc_subgroup_title"]&s=[{"cpc_subgroup_id":"asc"}]&o={"matched_subentities_only":true}

# example 2
# https://www.patentsview.org/api/cpc_subsections/query?q={"_and":[{"examiner_last_name":"fitzhugh","examiner_first_name":"julius"}]}&f=["cpc_subgroup_id","cpc_subgroup_title"]&s=[{"cpc_subgroup_id":"asc"}]&o={"matched_subentities_only":true}

# https://www.patentsview.org/api/cpc_subsections/query?
# q={"cpc_subgroup_id":"B32B1/06"}
# q={"_and":[{"examiner_last_name":"fitzhugh","examiner_first_name":"julius"}
# f=["cpc_subgroup_id","cpc_subgroup_title"]
# s=[{"cpc_subgroup_id":"asc"}]
# o={"matched_subentities_only":true}

PATENTVIEW = 'https://www.patentsview.org/api/cpc_subsections/query'
# QUERY_SYBMOL = "B32B1/06"
batch_query = ["B32B1/06", "A01G31/06", "A22C13/0013"]
# TODO: possibly turn this into a stream, so the limit does not matter
payload = {
  "q": {"cpc_subgroup_id": examiner.symbols_list_only[0:200]},
  "f": ["cpc_subgroup_id", "cpc_subgroup_title"],
  "s": [{"cpc_subgroup_id": "asc"}],
  "o": {"matched_subentities_only": 'true'}
}
# print(examiner.symbols_list_only[0:20])

# req = requests.get(PATENTVIEW, params=params)
# print(req.status_code)
# print(req.json()['cpc_subsections'][0]['cpc_subgroups'][0]['cpc_subgroup_id'])
# print(req.json()['cpc_subsections'][0]['cpc_subgroups'][0]['cpc_subgroup_title'])
req = requests.post(PATENTVIEW, data=json.dumps(payload))
if req.status_code != 200:
    print(req.headers['x-status-reason'])
    print(req.raise_for_status())    
    # print(req.http_error_msg)    
# print(req.json())
for subsection in req.json()['cpc_subsections']:
    for subgroup in subsection['cpc_subgroups']:
        subgroup_id = subgroup['cpc_subgroup_id']
        subgroup_title = subgroup['cpc_subgroup_title']
        # print(examiner.symbols[subgroup_id].symbol)
        examiner.symbols[subgroup_id].title = subgroup_title
        # print(subgroup_id)
        # print(type(examiner.symbols[subgroup_id]['title']))
        # print(examiner.symbols[subgroup_title]['title'])
        # print(subgroup['cpc_subgroup_id'])
        # print(subgroup['cpc_subgroup_title'])
# print(req.json()['cpc_subsections'][0]['cpc_subgroups'][0]['cpc_subgroup_id'])
# print(req.json()['cpc_subsections'][0]['cpc_subgroups'][0]['cpc_subgroup_title'])

# TODO: check status codes
# 200 ok
# 400 not valid query
# 500 internal error


# make new workbook and send to user
new_wb = Workbook()
# ws = new_wb.create_sheet('Processed Portfolio')
ws = new_wb.active
ws.page_setup.fitToWidth
ws.title = 'Processed Portfolio'
# <Symbol, title, tallies, c_start, qualified>
column_labels = ['Symbol', 'Title', 'Tally', 'C*', 'Qualified']
max_rows = len(examiner.symbols_list_only)

cell_range = ws['A1:E1']
max_columns = len(column_labels)

# for labelcell in cell_range[0]:
for i in range(max_columns):
    cell_range[0][i].value = column_labels[i]
    print(cell_range[0][i].value)

LABEL = 1
SYMBOLS_START = 2
SYMBOLS_END = max_rows

# print(examiner.portfolio)
# for row in range(2, 20):
for symbol in examiner.portfolio:
    # print(symbol)
    # for i in range(max_columns):
        # print(examiner.portfolio)
    # row_data = [examiner.portfolio[i].symbol,
                # examiner.portfolio[i].title,...
    row_data = [symbol.symbol,
                symbol.title,
                symbol.tally,
                symbol.c_star,
                symbol.qualified
                ]
    # print(row_data)
    ws.append(row_data)
# for cell in cell_range:
#     for label in column_labels:
#         print(cell)
#         print(label)

# for column in ws.iter_cols(max_col=5, max_row=1):
#     for label in column_labels:
#         print(column)
#         print(label)
        # column = label

# for row in ws.iter_rows(min_row=2, max_row={max_rows}, max_col=5):
#     pass

symbol_cells = ws['A']
title_cells = ws['B']
c_star_cells = ws['C']
tally_cells = ws['D']
qualified_cells = ws['E']
styles = {
  "align": Alignment(vertical='center', wrap_text=True),
  "align-center": Alignment(horizontal='center', vertical='center', wrap_text=True),
  "font": Font(size=16)
}


for cell in symbol_cells:
    cell.alignment = styles['align']
    cell.font = styles['font']
for cell in title_cells:
    cell.alignment = styles['align']
    cell.font = styles['font']
for cell in c_star_cells:
    cell.alignment = styles['align-center']
    cell.font = styles['font']
for cell in tally_cells:
    cell.alignment = styles['align-center']
    cell.font = styles['font']
for cell in qualified_cells:
    cell.alignment = styles['align-center']
    cell.font = styles['font']


new_wb.save(f'test_{PORTFOLIO}')
